"""M3 acceptance: all-or-nothing Excel import, exact error reporting,
CIN-idempotent re-import, MCA header mapping, audit diffs."""
from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook

from tests.conftest import csrf, register_firm

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
HEADERS = ["cin", "name", "agm_date", "paidup_capital", "fy_end_month"]


def cin(i: int) -> str:
    return f"U74999MH2020PTC{i:06d}"


def build_xlsx(rows: list[list], headers: list[str] = HEADERS) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def upload(client, data: bytes, dry_run: bool = False):
    return await client.post(
        f"/api/v1/companies/import?dry_run={str(dry_run).lower()}",
        files={"file": ("portfolio.xlsx", data, XLSX)},
        headers={"X-CSRF-Token": await csrf(client)},
    )


async def test_200_rows_with_5_bad_imports_nothing_reports_exactly_5(make_client) -> None:
    client = make_client()
    await register_firm(client, "import@example.com")

    rows = [[cin(i), f"Company {i}", "30/09/2026", 100000, 3] for i in range(195)]
    bad = [
        ["SHORT-CIN", "Bad CIN Co", "30/09/2026", 100000, 3],          # row 197
        [cin(900), "", "30/09/2026", 100000, 3],                       # row 198: no name
        [cin(901), "Bad Date Co", "not-a-date", 100000, 3],            # row 199
        [cin(902), "Bad Capital Co", "30/09/2026", "lots", 3],         # row 200
        [cin(0), "Duplicate CIN Co", "30/09/2026", 100000, 3],         # row 201: dup of row 2
    ]
    res = await upload(client, build_xlsx(rows + bad))
    assert res.status_code == 422
    errors = res.json()["detail"]["errors"]
    assert sorted({e["row"] for e in errors}) == [197, 198, 199, 200, 201]

    # NOTHING was imported (atomicity)
    listing = await client.get("/api/v1/companies")
    assert listing.headers["X-Total-Count"] == "0"


async def test_valid_import_commits_all(make_client) -> None:
    client = make_client()
    await register_firm(client, "import2@example.com")
    res = await upload(client, build_xlsx([[cin(i), f"Co {i}", None, 500000, 3] for i in range(10)]))
    assert res.status_code == 200, res.text
    assert res.json()["created"] == 10
    listing = await client.get("/api/v1/companies")
    assert listing.headers["X-Total-Count"] == "10"


async def test_dry_run_commits_nothing(make_client) -> None:
    client = make_client()
    await register_firm(client, "dry@example.com")
    res = await upload(client, build_xlsx([[cin(1), "Dry Co", None, None, None]]), dry_run=True)
    assert res.status_code == 200
    assert res.json() == {"rows_ok": 1, "errors": [], "dry_run": True, "imported": False}
    assert (await client.get("/api/v1/companies")).headers["X-Total-Count"] == "0"


async def test_reimport_idempotent_on_cin(make_client) -> None:
    client = make_client()
    await register_firm(client, "idem@example.com")
    file1 = build_xlsx([[cin(i), f"Co {i}", None, 500000, 3] for i in range(5)])
    assert (await upload(client, file1)).json()["created"] == 5

    # identical re-import: nothing created, nothing updated
    again = (await upload(client, file1)).json()
    assert (again["created"], again["updated"], again["unchanged"]) == (0, 0, 5)

    # one changed name → exactly one update, with an audit diff
    file2 = build_xlsx(
        [[cin(0), "Renamed Co", None, 500000, 3]]
        + [[cin(i), f"Co {i}", None, 500000, 3] for i in range(1, 5)]
    )
    changed = (await upload(client, file2)).json()
    assert (changed["created"], changed["updated"], changed["unchanged"]) == (0, 1, 4)

    import app.db as dbmod
    from sqlalchemy import select

    from app.models import ActivityLog

    async with dbmod._sessionmaker() as session:  # type: ignore[union-attr]
        entry = (
            await session.execute(
                select(ActivityLog).where(ActivityLog.action == "import_update")
            )
        ).scalars().one()
    assert entry.diff["before"]["name"] == "Co 0"
    assert entry.diff["after"]["name"] == "Renamed Co"


async def test_mca_master_data_headers_accepted(make_client) -> None:
    client = make_client()
    await register_firm(client, "mca@example.com")
    data = build_xlsx(
        [[cin(7), "MCA Import Co", "01/04/2019", "1,000,000", "ACTIVE"]],
        headers=["CIN", "COMPANY NAME", "DATE OF INCORPORATION", "PAID UP CAPITAL",
                 "COMPANY STATUS"],
    )
    res = await upload(client, data)
    assert res.status_code == 200, res.text
    assert res.json()["created"] == 1
    company = (await client.get("/api/v1/companies")).json()[0]
    assert company["incorporation_date"] == "2019-04-01"
    assert company["status"] == "ACTIVE"


async def test_viewer_cannot_import(firm) -> None:
    res = await upload(firm.viewer, build_xlsx([[cin(1), "Denied Co", None, None, None]]))
    assert res.status_code == 403


async def test_template_and_export_roundtrip(make_client) -> None:
    client = make_client()
    await register_firm(client, "roundtrip@example.com")

    template = await client.get("/api/v1/companies/import/template")
    assert template.status_code == 200
    ws = load_workbook(io.BytesIO(template.content)).active
    assert [c.value for c in ws[1]][:2] == ["cin", "name"]

    await upload(client, build_xlsx([[cin(3), "Export Co", "30/09/2026", 750000, 3]]))
    export = await client.get("/api/v1/companies/export")
    assert export.status_code == 200
    rows = list(load_workbook(io.BytesIO(export.content)).active.values)
    assert rows[1][0] == cin(3)
    assert rows[1][1] == "Export Co"


# ---- M15: per-master import/export (PRD §4.3/4.4) ----

MASTER_COMPANY = {"cin": "U74999MH2020PTC454545", "name": "Master Import Co"}


async def master_upload(client, cid: str, master: str, data: bytes, dry_run: bool = False):
    return await client.post(
        f"/api/v1/companies/{cid}/{master}/import?dry_run={str(dry_run).lower()}",
        files={"file": (f"{master}.xlsx", data, XLSX)},
        headers={"X-CSRF-Token": await csrf(client)},
    )


async def test_directors_import_all_or_nothing_and_idempotent(firm) -> None:
    from tests.conftest import post as _post

    cid = (await _post(firm.manager, "/companies", MASTER_COMPANY)).json()["id"]
    good = [
        ["A. Director", "01234567", "approved", "2015-01-01", "MD", "2020-06-01", None],
        ["B. Member", "07654321", None, None, "Director", None, None],
    ]
    headers = ["name", "din", "din_status", "din_allocation_date", "designation",
               "appointment_date", "cessation_date"]

    # a bad DIN poisons the whole file — nothing imports
    res = await master_upload(firm.executive, cid, "directors",
                              build_xlsx(good + [["C. Bad", "123", None, None, None, None, None]],
                                         headers=headers))
    assert res.status_code == 422
    assert res.json()["detail"]["errors"][0]["column"] == "din"
    assert (await firm.viewer.get(f"/api/v1/companies/{cid}/directors")).json() == []

    # clean file imports; identical re-import skips everything
    res = await master_upload(firm.executive, cid, "directors", build_xlsx(good, headers=headers))
    assert res.status_code == 200, res.text
    assert res.json()["created"] == 2
    res = await master_upload(firm.executive, cid, "directors", build_xlsx(good, headers=headers))
    assert (res.json()["created"], res.json()["skipped"]) == (0, 2)

    # export round-trip carries the data
    export = await firm.viewer.get(f"/api/v1/companies/{cid}/directors/export")
    rows = list(load_workbook(io.BytesIO(export.content)).active.values)
    assert rows[1][0] == "A. Director" and rows[1][1] == "01234567"


async def test_shareholders_import_and_rbac(firm) -> None:
    from tests.conftest import post as _post

    cid = (await _post(firm.manager, "/companies", MASTER_COMPANY)).json()["id"]
    headers = ["name", "folio", "shares", "percentage", "category"]
    data = build_xlsx([["Holder One", "F001", 9000, 90, "promoter"],
                       ["Holder Two", "F002", 1000, 10, "public"]], headers=headers)

    assert (await master_upload(firm.viewer, cid, "shareholders", data)).status_code == 403

    res = await master_upload(firm.executive, cid, "shareholders", data)
    assert res.status_code == 200, res.text
    assert res.json()["created"] == 2
    holders = (await firm.viewer.get(f"/api/v1/companies/{cid}/shareholders")).json()
    assert holders["total_shares"] == "10000"

    # unknown master name → 404, not a crash
    res = await master_upload(firm.executive, cid, "charges", data)
    assert res.status_code == 404
