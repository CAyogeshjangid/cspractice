"""Company Excel import/export (charter M3).

Contract:
- Validation is ALL-OR-NOTHING: any row error → nothing imports, and the
  report lists exactly the failing rows/columns (M3 acceptance).
- Re-import is idempotent on CIN: unchanged rows are skipped, changed rows
  are updated with an audit diff, soft-deleted rows are restored (REVIEW F8).
- Our template headers AND common MCA master-data headers are both accepted.
- No dates are computed here — the calendar belongs to the rules engine (C12).
"""
from __future__ import annotations

import io
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from starlette.requests import Request

from app import audit
from app.models import Company
from app.repositories import companies as companies_repo

TEMPLATE_HEADERS = [
    "cin", "name", "registration_number", "incorporation_date", "category",
    "status", "registered_address", "email", "phone", "fy_end_month",
    "fy_end_day", "agm_date", "is_listed", "authorised_capital",
    "subscribed_capital", "paidup_capital",
]

# Common MCA master-data export headers → our fields (case/space-insensitive).
MCA_HEADER_MAP = {
    "CIN": "cin",
    "COMPANY NAME": "name",
    "COMPANY_NAME": "name",
    "REGISTRATION NUMBER": "registration_number",
    "REGISTRATION_NUM": "registration_number",
    "DATE OF INCORPORATION": "incorporation_date",
    "DATE_OF_REGISTRATION": "incorporation_date",
    "COMPANY CATEGORY": "category",
    "COMPANY_CATEGORY": "category",
    "COMPANY STATUS": "status",
    "COMPANY_STATUS": "status",
    "REGISTERED ADDRESS": "registered_address",
    "REGISTERED_OFFICE_ADDRESS": "registered_address",
    "EMAIL": "email",
    "EMAIL_ADDR": "email",
    "AUTHORIZED CAPITAL": "authorised_capital",
    "AUTHORIZED_CAP": "authorised_capital",
    "AUTHORISED CAPITAL": "authorised_capital",
    "PAID UP CAPITAL": "paidup_capital",
    "PAIDUP CAPITAL": "paidup_capital",
    "PAIDUP_CAPITAL": "paidup_capital",
}

DATE_FIELDS = {"incorporation_date", "agm_date"}
DECIMAL_FIELDS = {"authorised_capital", "subscribed_capital", "paidup_capital"}
INT_FIELDS = {"fy_end_month", "fy_end_day"}
BOOL_FIELDS = {"is_listed"}


@dataclass
class RowError:
    row: int  # 1-based spreadsheet row number
    column: str
    error: str

    def as_dict(self) -> dict[str, Any]:
        return {"row": self.row, "column": self.column, "error": self.error}


@dataclass
class ParseResult:
    rows: list[dict[str, Any]] = field(default_factory=list)  # includes _row marker
    errors: list[RowError] = field(default_factory=list)


def build_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "companies"
    ws.append(TEMPLATE_HEADERS)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _map_header(raw: str) -> str | None:
    cleaned = str(raw).strip()
    if cleaned.lower() in TEMPLATE_HEADERS:
        return cleaned.lower()
    return MCA_HEADER_MAP.get(cleaned.upper())


def _parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"unparseable date: {text!r} (use DD/MM/YYYY)")


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "yes", "y", "1"}


def parse_and_validate(data: bytes) -> ParseResult:
    result = ParseResult()
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        result.errors.append(RowError(0, "-", "file is not a readable .xlsx workbook"))
        return result

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        result.errors.append(RowError(0, "-", "workbook is empty"))
        return result

    columns: dict[int, str] = {}
    for idx, raw in enumerate(header_row):
        if raw is None:
            continue
        mapped = _map_header(str(raw))
        if mapped:
            columns[idx] = mapped
    if "cin" not in columns.values() or "name" not in columns.values():
        result.errors.append(
            RowError(1, "-", "header row must include at least CIN and company name columns")
        )
        return result

    seen_cins: dict[str, int] = {}
    for row_no, values in enumerate(rows_iter, start=2):
        if values is None or all(v is None or str(v).strip() == "" for v in values):
            continue  # blank line
        record: dict[str, Any] = {"_row": row_no}
        row_errors: list[RowError] = []
        for idx, field_name in columns.items():
            value = values[idx] if idx < len(values) else None
            if value is None or str(value).strip() == "":
                continue
            try:
                if field_name in DATE_FIELDS:
                    record[field_name] = _parse_date(value)
                elif field_name in DECIMAL_FIELDS:
                    record[field_name] = Decimal(str(value).replace(",", "").strip())
                elif field_name in INT_FIELDS:
                    record[field_name] = int(str(value).strip())
                elif field_name in BOOL_FIELDS:
                    record[field_name] = _parse_bool(value)
                else:
                    record[field_name] = str(value).strip()
            except (ValueError, InvalidOperation) as exc:
                row_errors.append(RowError(row_no, field_name, str(exc)))

        cin = record.get("cin", "")
        if len(cin) != 21:
            row_errors.append(RowError(row_no, "cin", f"CIN must be 21 characters, got {len(cin)}"))
        elif cin in seen_cins:
            row_errors.append(
                RowError(row_no, "cin", f"duplicate CIN in file (first seen at row {seen_cins[cin]})")
            )
        else:
            seen_cins[cin] = row_no
        if not record.get("name"):
            row_errors.append(RowError(row_no, "name", "company name is required"))
        month, day = record.get("fy_end_month"), record.get("fy_end_day")
        if month is not None and not 1 <= month <= 12:
            row_errors.append(RowError(row_no, "fy_end_month", "must be 1–12"))
        if day is not None and not 1 <= day <= 31:
            row_errors.append(RowError(row_no, "fy_end_day", "must be 1–31"))

        result.errors.extend(row_errors)
        if not row_errors:
            result.rows.append(record)
    return result


COMPANY_FIELDS = [h for h in TEMPLATE_HEADERS]


async def commit_rows(
    session: AsyncSession,
    firm_id: uuid.UUID,
    actor_user_id: uuid.UUID,
    rows: list[dict[str, Any]],
    request: Request | None = None,
) -> dict[str, int]:
    """Atomic upsert-by-CIN. Caller guarantees rows validated (no errors)."""
    created = updated = restored = unchanged = 0
    for record in rows:
        data = {k: v for k, v in record.items() if k != "_row"}
        existing = await companies_repo.get_by_cin(session, firm_id, data["cin"])
        if existing is None:
            company = await companies_repo.create_company(session, firm_id, data)
            await audit.record(
                session, firm_id=firm_id, actor_user_id=actor_user_id, entity_type="company",
                entity_id=company.id, action="import_create",
                after={k: str(v) for k, v in data.items()}, request=request,
            )
            created += 1
            continue

        was_deleted = existing.deleted_at is not None
        diff_before: dict[str, str] = {}
        diff_after: dict[str, str] = {}
        for key, value in data.items():
            current = getattr(existing, key)
            if current != value and not (current is None and value is None):
                diff_before[key] = str(current)
                diff_after[key] = str(value)
                setattr(existing, key, value)
        if was_deleted:
            diff_before["deleted_at"] = str(existing.deleted_at)
            existing.deleted_at = None
            existing.deleted_reason = None
            restored += 1
        elif diff_after:
            updated += 1
        else:
            unchanged += 1
            continue
        await session.flush()
        await audit.record(
            session, firm_id=firm_id, actor_user_id=actor_user_id, entity_type="company",
            entity_id=existing.id,
            action="import_restore" if was_deleted else "import_update",
            before=diff_before, after=diff_after, request=request,
        )
    await session.commit()
    return {"created": created, "updated": updated, "restored": restored, "unchanged": unchanged}


def build_export(companies: list[Company]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "companies"
    ws.append(TEMPLATE_HEADERS)
    for c in companies:
        ws.append([_cell(getattr(c, h)) for h in TEMPLATE_HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)  # display only; storage stays NUMERIC (§9)
    return value


# ---------------------------------------------------------------------------
# Generic per-master importer (M15): directors and shareholders get the same
# all-or-nothing contract as companies (PRD §4.3/4.4).

MASTER_SPECS: dict[str, dict[str, Any]] = {
    "directors": {
        "headers": ["name", "din", "din_status", "din_allocation_date", "designation",
                    "appointment_date", "cessation_date"],
        "required": ["name"],
        "dates": {"din_allocation_date", "appointment_date", "cessation_date"},
        "decimals": set(),
        # duplicate-skip identity on re-import (no natural unique key like CIN)
        "identity": lambda rec: (rec.get("din") or "", rec.get("name", "")),
    },
    "shareholders": {
        "headers": ["name", "folio", "shares", "percentage", "category"],
        "required": ["name"],
        "dates": set(),
        "decimals": {"shares", "percentage"},
        "identity": lambda rec: (rec.get("folio") or "", rec.get("name", "")),
    },
}


def build_master_template(master: str) -> bytes:
    spec = MASTER_SPECS[master]
    wb = Workbook()
    ws = wb.active
    ws.title = master
    ws.append(spec["headers"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_master(master: str, data: bytes) -> ParseResult:
    """Same contract as the company importer: any error → nothing imports."""
    spec = MASTER_SPECS[master]
    result = ParseResult()
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        result.errors.append(RowError(0, "-", "file is not a readable .xlsx workbook"))
        return result
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        result.errors.append(RowError(0, "-", "workbook is empty"))
        return result

    columns: dict[int, str] = {}
    for idx, raw in enumerate(header_row):
        if raw is not None and str(raw).strip().lower() in spec["headers"]:
            columns[idx] = str(raw).strip().lower()
    missing_headers = [h for h in spec["required"] if h not in columns.values()]
    if missing_headers:
        result.errors.append(RowError(1, "-", f"header row must include: {missing_headers}"))
        return result

    for row_no, values in enumerate(rows_iter, start=2):
        if values is None or all(v is None or str(v).strip() == "" for v in values):
            continue
        record: dict[str, Any] = {"_row": row_no}
        row_errors: list[RowError] = []
        for idx, field_name in columns.items():
            value = values[idx] if idx < len(values) else None
            if value is None or str(value).strip() == "":
                continue
            try:
                if field_name in spec["dates"]:
                    record[field_name] = _parse_date(value)
                elif field_name in spec["decimals"]:
                    record[field_name] = Decimal(str(value).replace(",", "").strip())
                else:
                    record[field_name] = str(value).strip()
            except (ValueError, InvalidOperation) as exc:
                row_errors.append(RowError(row_no, field_name, str(exc)))
        for field_name in spec["required"]:
            if not record.get(field_name):
                row_errors.append(RowError(row_no, field_name, f"{field_name} is required"))
        din = record.get("din")
        if master == "directors" and din and len(str(din)) != 8:
            row_errors.append(RowError(row_no, "din", "DIN must be 8 digits"))
        result.errors.extend(row_errors)
        if not row_errors:
            result.rows.append(record)
    return result
